"""
Data Cleaner & Report Tool
----------------------------
أداة بسيطة تاخذ ملف Excel أو CSV "خام" وتسوي له:
1. تنظيف تلقائي (حذف التكرار، تعبئة القيم الفارغة، توحيد التواريخ)
2. تقرير إحصائي سريع عن جودة البيانات
3. تصدير ملف نظيف + تقرير Excel بالنتائج

طريقة الاستخدام:
    python cleaner.py my_file.csv
    python cleaner.py my_file.xlsx
"""

import pandas as pd
import sys
import os
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- إعدادات التنسيق العام ---
FONT_NAME = "Calibri"
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name=FONT_NAME, size=10.5)
TITLE_FONT = Font(name=FONT_NAME, size=14, bold=True, color="1F4E78")
ALT_FILL = PatternFill(start_color="F2F6FA", end_color="F2F6FA", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)


def style_sheet(ws, has_header=True, zebra=True, freeze=True, right_to_left=True):
    """يطبّق تنسيق احترافي موحّد على أي شيت: عناوين ملونة، حدود،
    تلوين متبادل للصفوف، عرض أعمدة تلقائي، وتجميد الصف الأول."""
    ws.sheet_view.rightToLeft = right_to_left

    max_row = ws.max_row
    max_col = ws.max_column

    # تنسيق الصف الأول (العناوين)
    if has_header:
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER
            cell.border = THIN_BORDER
        ws.row_dimensions[1].height = 24

    # تنسيق باقي الصفوف
    start_row = 2 if has_header else 1
    for row_idx in range(start_row, max_row + 1):
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = RIGHT
            if zebra and (row_idx - start_row) % 2 == 1:
                cell.fill = ALT_FILL

    # عرض أعمدة تلقائي حسب أطول محتوى
    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 45)

    if freeze and has_header:
        ws.freeze_panes = "A2"


def load_file(filepath):
    """يقرأ الملف حسب امتداده (CSV أو Excel)"""
    if filepath.endswith(".csv"):
        return pd.read_csv(filepath)
    elif filepath.endswith((".xlsx", ".xls")):
        return pd.read_excel(filepath)
    else:
        raise ValueError("الصيغة غير مدعومة. استخدمي ملف .csv أو .xlsx")


def profile_data(df):
    """يسوي ملخص عن جودة البيانات قبل التنظيف"""
    profile = {
        "عدد الصفوف": len(df),
        "عدد الأعمدة": len(df.columns),
        "عدد الصفوف المكررة": df.duplicated().sum(),
        "إجمالي القيم الفارغة": df.isnull().sum().sum(),
    }

    # عدد القيم الفارغة لكل عمود
    missing_per_column = df.isnull().sum()
    missing_per_column = missing_per_column[missing_per_column > 0]

    return profile, missing_per_column


def clean_data(df):
    """ينظف البيانات: يحذف التكرار، يعالج القيم الفارغة، يوحد التواريخ"""
    df_clean = df.copy()

    # 1. حذف الصفوف المكررة بالكامل
    df_clean = df_clean.drop_duplicates()

    # 2. معالجة القيم الفارغة حسب نوع العمود
    for col in df_clean.columns:
        if df_clean[col].dtype in ["float64", "int64"]:
            # أعمدة رقمية: نعبي بالمتوسط، مقرّب لأقرب رقمين عشريين
            mean_val = round(df_clean[col].mean(), 2)
            df_clean[col] = df_clean[col].fillna(mean_val)
            # لو كل القيم أعداد صحيحة أصلاً، نخليها int بدل float
            if (df_clean[col] % 1 == 0).all():
                df_clean[col] = df_clean[col].astype(int)
        else:
            # أعمدة نصية: نعبي بكلمة "غير محدد"
            df_clean[col] = df_clean[col].fillna("غير محدد")

    # 3. محاولة توحيد أعمدة التاريخ (لو فيه عمود اسمه فيه "date" أو "تاريخ")
    for col in df_clean.columns:
        if "date" in col.lower() or "تاريخ" in col:
            try:
                df_clean[col] = pd.to_datetime(df_clean[col], errors="coerce").dt.date
            except Exception:
                pass

    # 4. إزالة المسافات الزائدة من الأعمدة النصية
    for col in df_clean.select_dtypes(include=["object", "string"]).columns:
        df_clean[col] = df_clean[col].astype(str).str.strip()

    return df_clean


def export_results(df_clean, profile_before, missing_per_column, output_name):
    """يصدر ملف نظيف + تقرير Excel بتنسيق احترافي"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")

    clean_filename = f"{output_name}_clean_{timestamp}.xlsx"
    report_filename = f"{output_name}_report_{timestamp}.xlsx"

    # ---------- 1) ملف البيانات النظيفة ----------
    with pd.ExcelWriter(clean_filename, engine="openpyxl") as writer:
        df_clean.to_excel(writer, sheet_name="البيانات النظيفة", index=False)
        style_sheet(writer.sheets["البيانات النظيفة"])

    # ---------- 2) ملف التقرير ----------
    with pd.ExcelWriter(report_filename, engine="openpyxl") as writer:
        # --- شيت الملخص العام ---
        summary_df = pd.DataFrame(
            list(profile_before.items()), columns=["المؤشر", "القيمة"]
        )
        summary_df.to_excel(writer, sheet_name="ملخص عام", index=False, startrow=2)
        ws_summary = writer.sheets["ملخص عام"]
        ws_summary["A1"] = f"تقرير جودة البيانات — {output_name}"
        ws_summary["A1"].font = TITLE_FONT
        ws_summary.merge_cells("A1:B1")
        ws_summary["A1"].alignment = RIGHT
        # نعيد التنسيق بعد الإزاحة (الهيدر الآن بالصف 3)
        _style_offset_sheet(ws_summary, header_row=3)

        # --- شيت القيم الفارغة لكل عمود ---
        if len(missing_per_column) > 0:
            missing_df = missing_per_column.reset_index()
            missing_df.columns = ["العمود", "عدد القيم الفارغة"]
            missing_df.to_excel(writer, sheet_name="قيم فارغة لكل عمود", index=False)
            style_sheet(writer.sheets["قيم فارغة لكل عمود"])

        # --- شيت الوصف الإحصائي ---
        describe_df = df_clean.describe(include="all").transpose().reset_index()
        describe_df.rename(columns={"index": "العمود"}, inplace=True)
        describe_df.to_excel(writer, sheet_name="وصف إحصائي", index=False)
        style_sheet(writer.sheets["وصف إحصائي"])

    return clean_filename, report_filename


def _style_offset_sheet(ws, header_row):
    """نسخة من style_sheet تدعم عناوين تبدأ من صف غير الأول (بسبب عنوان التقرير)."""
    ws.sheet_view.rightToLeft = True
    max_row = ws.max_row
    max_col = ws.max_column

    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    ws.row_dimensions[header_row].height = 24

    for row_idx in range(header_row + 1, max_row + 1):
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = RIGHT
            if (row_idx - header_row) % 2 == 0:
                cell.fill = ALT_FILL

    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 45)

    ws.freeze_panes = f"A{header_row + 1}"


def main():
    if len(sys.argv) < 2:
        print("الاستخدام: python cleaner.py اسم_الملف.csv")
        sys.exit(1)

    filepath = sys.argv[1]

    if not os.path.exists(filepath):
        print(f"الملف غير موجود: {filepath}")
        sys.exit(1)

    print(f"📂 جاري قراءة الملف: {filepath}")
    df = load_file(filepath)

    print("🔍 جاري تحليل جودة البيانات...")
    profile_before, missing_per_column = profile_data(df)

    print("\n--- ملخص قبل التنظيف ---")
    for key, value in profile_before.items():
        print(f"{key}: {value}")

    print("\n🧹 جاري تنظيف البيانات...")
    df_clean = clean_data(df)

    output_name = os.path.splitext(os.path.basename(filepath))[0]
    clean_file, report_file = export_results(
        df_clean, profile_before, missing_per_column, output_name
    )

    print(f"\n✅ تم! الملفات الناتجة:")
    print(f"   - ملف نظيف: {clean_file}")
    print(f"   - تقرير: {report_file}")


if __name__ == "__main__":
    main()
